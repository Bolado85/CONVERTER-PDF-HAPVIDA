"""
Exportador de dados para Excel com formatação profissional.
Gera 3 abas: Dados Completos, Por Especialidade, Por Cidade.
"""
import io
from typing import List, Optional
import pandas as pd


def exportar_excel(df: pd.DataFrame, cidades: List[str]) -> Optional[bytes]:
    try:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            colunas = [c for c in [
                "UF","CIDADE","BAIRRO","ENDEREÇO","NÚMERO","COMPLEMENTO",
                "PONTO REFERENCIA","TELEFONE","TELEFONE1",
                "NOME CLINICA VINCULADA","ESPECIALIDADE","TIPO PRESTADOR",
                "CREDENCIADO","NOME FANTASIA","CNPJ_CPF"
            ] if c in df.columns]
            df_out = df[colunas].sort_values(["CIDADE","ESPECIALIDADE","NOME FANTASIA"], na_position="last")
            df_out.to_excel(writer, sheet_name="Dados Completos", index=False)
            ws1 = writer.sheets["Dados Completos"]
            _formatar(ws1)

            res_esp = (df.groupby(["UF","CIDADE","ESPECIALIDADE"]).size()
                       .reset_index(name="Prestadores")
                       .sort_values(["CIDADE","ESPECIALIDADE"]))
            res_esp.to_excel(writer, sheet_name="Por Especialidade", index=False)
            _formatar(writer.sheets["Por Especialidade"])

            res_cid = (df.groupby(["UF","CIDADE"])
                       .agg(Prestadores=("NOME FANTASIA","count"),
                            Especialidades=("ESPECIALIDADE","nunique"))
                       .reset_index().sort_values("CIDADE"))
            res_cid.to_excel(writer, sheet_name="Por Cidade", index=False)
            _formatar(writer.sheets["Por Cidade"])

        return buffer.getvalue()
    except Exception:
        import traceback; traceback.print_exc()
        return None


def _formatar(ws):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        fill_h = PatternFill(start_color="002B5C", end_color="002B5C", fill_type="solid")
        font_h = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        fill_p = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        font_n = Font(size=9, name="Calibri")
        borda  = Border(bottom=Side(style="hair", color="DDE3EC"))

        for cell in ws[1]:
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"

        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.font = font_n
                cell.border = borda
                if i % 2 == 0:
                    cell.fill = fill_p

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)
    except Exception:
        pass